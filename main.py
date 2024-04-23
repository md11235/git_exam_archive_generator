# This is a sample Python script.
import argparse

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from openpyxl.styles import Font, Color, Border, Side
import os
import argparse


DAILY_PERF = "平时成绩"
LABS = "课堂实验"
LECTURE_PERF = "课堂表现"

CLASS_NAME = "班级"

ATTENDANCE = "考勤"

NO_OF_ATTENDANCE = 10


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def extract_daily_performance(all_in_one_workbook_file, class_name, class_info_file,
                              output_template_workbook="05平时成绩模板.xlsx",
                              output_start_row=5,
                              num_attendances=10):
    df_daily_perf = pd.read_excel(all_in_one_workbook_file, sheet_name=DAILY_PERF, header=0)

    df_labs = pd.read_excel(all_in_one_workbook_file, sheet_name=LABS, header=0)
    print(df_labs)

    df_lecture_perf = pd.read_excel(all_in_one_workbook_file, sheet_name=LECTURE_PERF, header=0)
    print(df_lecture_perf)

    df_stu_info = pd.read_csv(class_info_file, sep='\t', header=None)

    print(df_stu_info)

    df_index = df_stu_info[[0, 1]].rename({0: "学号", 1: "姓名"}, axis=1)
    # for att_index in range(1, NO_OF_ATTENDANCE+1):
    #     df_result[ATTENDANCE+str(att_index)] = df_daily_perf
    # print(df_result)

    print(df_index["学号"])
    df_daily_perf.set_index("学号", inplace=True)
    df_index.set_index("学号", inplace=True)
    # print(df_daily_perf)

    # print(pd.concat([df_result, df_daily_perf.loc[df_result.index]], axis=1))
    print("--------")
    df_result = df_daily_perf.loc[df_index.index].reset_index()

    script_dir = os.path.dirname(os.path.realpath(__file__))
    template_wb = openpyxl.load_workbook(os.path.join(script_dir, output_template_workbook))

    template_ws = None
    if DAILY_PERF in template_wb:
        template_ws = template_wb[DAILY_PERF]
    else:
        raise ValueError("Can't find the template sheet for {}".format(DAILY_PERF))

    df_result[CLASS_NAME] = class_name
    print(df_result)

    cell_font = Font(name="SimSun", size=11)

    for ind, row in df_result.iterrows():
        row_id = output_start_row+ind
        template_ws["A{}".format(row_id)] = str(row["学号"])
        template_ws["A{}".format(row_id)].font = cell_font
        template_ws["A{}".format(row_id)].number_format = numbers.FORMAT_TEXT

        template_ws["C{}".format(row_id)] = row["姓名"]
        template_ws["C{}".format(row_id)].font = cell_font

        template_ws["E{}".format(row_id)] = row["班级"]
        template_ws["E{}".format(row_id)].font = cell_font

        template_ws["{}{}".format('I', row_id)] = row["考勤{}".format(1)]
        attendance_first_column_ord = ord("K")
        for i in range(1, num_attendances):
            template_ws["{}{}".format(chr(attendance_first_column_ord+i), row_id)] = row["考勤{}".format(i+1)]

        template_ws["{}{}".format(chr(attendance_first_column_ord+num_attendances), row_id)] = row["考勤分"]
        template_ws["{}{}".format(chr(attendance_first_column_ord+num_attendances+1), row_id)] = row["实验"]
        template_ws["{}{}".format(chr(attendance_first_column_ord+num_attendances+2), row_id)] = row["课堂表现"]
        template_ws["{}{}".format(chr(attendance_first_column_ord+num_attendances+3), row_id)] = row["总成绩"]

        final_column = chr(attendance_first_column_ord+num_attendances+3)

        # TODO: merge E F G H

        set_border(template_ws, "A{}:{}{}".format(row_id, final_column, row_id))

    # for row in dataframe_to_rows(df_result, index=False, header=False):
    #     print(row)
    #     # template_ws.append(row)

    template_wb.save("05_gen_平时成绩_{}.xlsx".format(class_name))


def extract_class_performance(all_in_one_workbook_file, class_name, class_info_file,
                              output_template_workbook="05平时成绩模板.xlsx",
                              output_start_row=4,
                              num_chapters=10):
    df_lecture_perf = pd.read_excel(all_in_one_workbook_file, sheet_name=LECTURE_PERF, header=0)
    print(df_lecture_perf)

    df_stu_info = pd.read_csv(class_info_file, sep='\t', header=None)

    print(df_stu_info)

    df_index = df_stu_info[[0, 1]].rename({0: "学号", 1: "姓名"}, axis=1)
    # for att_index in range(1, NO_OF_ATTENDANCE+1):
    #     df_result[ATTENDANCE+str(att_index)] = df_daily_perf
    # print(df_result)

    print(df_index["学号"])
    df_lecture_perf.set_index("学号", inplace=True)
    df_index.set_index("学号", inplace=True)

    print("--------")
    df_result = df_lecture_perf.loc[df_index.index].reset_index()

    script_dir = os.path.dirname(os.path.realpath(__file__))
    template_wb = openpyxl.load_workbook(os.path.join(script_dir, output_template_workbook))

    template_ws = None
    if LECTURE_PERF in template_wb:
        template_ws = template_wb[LECTURE_PERF]
    else:
        raise ValueError("Can't find the template sheet for {}".format(LECTURE_PERF))

    df_result[CLASS_NAME] = class_name
    print(df_result)

    cell_font = Font(name="SimSun", size=11)

    chap_id2name = {14: "第15章",
                    15: "第17章"}

    for ind, row in df_result.iterrows():
        row_id = output_start_row + ind
        template_ws["A{}".format(row_id)] = ind+1
        template_ws["A{}".format(row_id)].font = cell_font

        template_ws["B{}".format(row_id)] = str(row["学号"])
        template_ws["B{}".format(row_id)].font = cell_font
        template_ws["B{}".format(row_id)].number_format = numbers.FORMAT_TEXT

        template_ws["C{}".format(row_id)] = row["姓名"]
        template_ws["C{}".format(row_id)].font = cell_font

        # template_ws["D{}".format(row_id)] = row["第1章"]
        # template_ws["D{}".format(row_id)].font = cell_font

        chapter_first_col_ord = ord('D')
        for i in range(num_chapters):
            chap_name = chap_id2name[i+1] if (i+1) in chap_id2name else "第{}章".format(i+1)
            template_ws["{}{}".format(chr(chapter_first_col_ord+i), row_id)] = row[chap_name]

        template_ws["{}{}".format(chr(chapter_first_col_ord+num_chapters), row_id)] = row["总分"]
        final_col = chr(chapter_first_col_ord+num_chapters+1)

        set_border(template_ws, "A{}:{}{}".format(row_id, final_col, row_id))

    # for row in dataframe_to_rows(df_result, index=False, header=False):
    #     print(row)
    #     # template_ws.append(row)

    template_wb.save("05_gen_课堂表现_{}.xlsx".format(class_name))


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("all_in_one_scores_file",
                            help="包含{}、{}、{}工作表的excel工作簿".format(DAILY_PERF, LABS, LECTURE_PERF))
    arg_parser.add_argument("class_name",
                            help="班级名字")
    arg_parser.add_argument("class_students_id_name_filepath",
                            help="包含班级里学生的学号、姓名的tsv文件路径")

    arg_parser.add_argument("-t", help="输出模板", dest="output_template_filepath")

    args = arg_parser.parse_args()

    extract_daily_performance(args.all_in_one_scores_file,
                              args.class_name,
                              args.class_students_id_name_filepath,
                              output_template_workbook=args.output_template_filepath,
                              num_attendances=9)

    extract_class_performance(args.all_in_one_scores_file,
                              args.class_name,
                              args.class_students_id_name_filepath,
                              output_template_workbook=args.output_template_filepath,
                              num_chapters=15)


    # extract_daily_performance("../05-平时成绩-all3classes_v2_归档.xlsx",
    #                           "大数据213",
    #                           "../学生名单_大数据213.csv")
    # extract_daily_performance("../05-平时成绩-all3classes_v2_归档.xlsx",
    #                           "网络211",
    #                           "../学生名单_网络211.csv")
    #
    # extract_daily_performance("../05-平时成绩-all3classes_v2_归档.xlsx",
    #                           "网络212",
    #                           "../学生名单_网络212.csv")
    # extract_class_performance("../05-平时成绩-all3classes_v2_归档.xlsx",
    #                           "大数据213",
    #                           "../学生名单_大数据213.csv")
    #
    # extract_class_performance("../05-平时成绩-all3classes_v2_归档.xlsx",
    #                       "网络211",
    #                       "../学生名单_网络211.csv")
    # extract_class_performance("../05-平时成绩-all3classes_v2_归档.xlsx",
    #                       "网络212",
    #                       "../学生名单_网络212.csv")

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
